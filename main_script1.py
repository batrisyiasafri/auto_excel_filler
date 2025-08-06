from openpyxl import Workbook

def main_script1(input_file, output_file):
    print(f"[DEBUG] Starting conversion: {input_file} -> {output_file}")
    try:
        with open(input_file, 'r') as file:
            lines = [line.strip() for line in file if line.strip()]
        print(f"[DEBUG] Read {len(lines)} lines from {input_file}")

        # wb = Workbook()
        # ws = wb.active

        records = []
        current_record = {}

        for line in lines:
            if ':' in line:
                key, value = line.split(':', 1)
                current_record[key.strip().capitalize()] = value.strip()
            elif not line.strip():
                if current_record:
                    records.append(current_record)
                    current_record = {}


        if current_record: #if last record wasnt added
                records.append(current_record)

           
        #for headers - alphabethcically due to sorted()
        # headers = sorted(set().union(*[r.keys() for r in records]))

        #headers - in the order they appear
        headers = []
        for record in records:
            for key in record:
                if key not in headers:
                    headers.append(key)
        print(f"[DEBUG] Headers detected: {headers}")

        wb = Workbook()
        ws = wb.active
        ws.append(headers)

        for record in records:
            row = [record.get(header, "") for header in headers]
            ws.append(row)
            print(f"[DEBUG] Writing row: {row}")

        wb.save(output_file)
        print(f"[SUCCESS] Excel file saved as {output_file}")

    except FileNotFoundError:
        print(f"[ERROR] The file '{input_file}' does not exist.")
    except PermissionError:
        print(f"[ERROR] Permission denied writing '{output_file}'.")
    except Exception as e:
        print(f"[ERROR] Unexpected error: {e}")





#to generate into cell row
    #  for row_idx, line in enumerate(lines, start=1):
    #         values = line.split(',') #split line by commas
    #         print(f"[DEBUG] Writing row {row_idx}: {values}")
    #         for col_idx, value in enumerate(values, start=1):
    #             ws.cell(row=row_idx, column=col_idx, value=value.strip())
            # ws[f'A{idx}'] = line

        #add header based on wat ur enter
        # ws.append(["Name," "Email", "Address", "Phone"])
#generate into 4 separate fields
        # row=[]
        # for idx, line in enumerate(lines, start=1):
        #     if ':' in line:
        #         key, value = line.split(':', 1)
        #         row.append(value.strip())
        #         #write to excel after 4 fields
        #         if len(row) == 4:
        #             ws.append(row)
        #             print(f"[DEBUG] Writing row: {row}")
        #             row = []
        
                    #detect new record on blank line if used
            #    else if line == '':
            #         records.append(current_record)
            #         current_record = []

                 #after complete record, supposed to be save n reset
            # if len(current_record) >= 3:
            #     records.append(current_record)
            #     current_record = {}